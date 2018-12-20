from openpyxl import load_workbook as load

dir = 'D:/FastCampus/FastCampusWebPythonBasic-python/code09_Excel/users.xlsx'
wb = load(dir)
#ws = wb.create_sheet('test')
ws = wb['test']
#ws['A1'] = '제목1'
#ws['B1'] = '제목2'
#
#wb.save(dir)
#wb.close()#

a1=ws['A1'].value
a2=ws['B1'].value

print(a1,a2)

wb.close()