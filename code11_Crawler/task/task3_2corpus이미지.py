from konlpy.tag import Okt
import matplotlib as mpl
mpl.use('TkAgg')
import matplotlib.font_manager as fm
from matplotlib import pyplot as plt
from wordcloud import WordCloud
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook as load
# import 구문이 추가로 필요하다는 점에 주의하세요!

URL = "http://www.ndsl.kr/ndsl/search/list/article/articleSearchResultList.do?page={}&query=%ED%8C%8C%EC%9D%B4%EC%8D%AC&prefixQuery=&collectionQuery=&showQuery=%ED%8C%8C%EC%9D%B4%EC%8D%AC&resultCount=10&sortName=RANK&sortOrder=DESC&colType=scholar&colTypeByUser=&filterValue="
PATH = '/Library/Fonts/NanumGothic.ttf'

def get_wc(content, title):
    ok = Okt()
    result = {}
    for c in content:
        temp = ok.pos(c)
        for t in temp:
            if t[1] == 'Noun' and len(t[0]) > 1:
                if not (t[0] in result):
                    result[t[0]] = 0
                result[t[0]] += 1
    result = sorted(result.items(), key = lambda x:x[1], reverse = True)
    print(dict(result))
    wc = WordCloud(font_path = PATH, background_color = 'white', width = 800, height = 600)
    plt.axis('off')
    plt.imshow(wc.generate_from_frequencies(dict(result)))
    plt.savefig(title + '.jpg')


def save_excel(a_links, b_links, title = 'test'):
    wb = load('test.xlsx')
    ws = wb.create_sheet(title)
    try:
        for i in range(len(a_links)):
            ws['A' + str(i + 1)] = a_links[i]
        for j in range(len(b_links)):
            ws['B' + str(j + 1)] = b_links[j]
        wb.save('test.xlsx')
    except Exception as e:
        print(e)
    finally:
        wb.close()

def get_link():
    a_links = []
    b_links = [] 
    for page in range(1, 11):
        req = requests.get(URL.format(str(page)))
        code = req.status_code
        print(code)
        if code == 200:
            soup = bs(req.text, 'html.parser')
            div = soup.find_all('div', attrs= {'class':'list_con'})
            for data in div:
                a_tag = data.find('a', attrs={'title': '상세화면'})
                a_tag = a_tag.get_text()        
                a_tag = a_tag.replace('\n', '').replace('\t', '')
                box_st1 = data.find('div', attrs = {'class':'box_st1'})
                box_st1 = box_st1.get_text()
                a_links.append(a_tag)
                b_links.append(box_st1)
        else:
            print('HTTP Error:', code)
    return a_links, b_links

if __name__ == "__main__":
    a_links, b_links = get_link()
    save_excel(a_links, b_links, title='content')
    get_wc(a_links, 'python제목')
    get_wc(b_links, 'python초록')