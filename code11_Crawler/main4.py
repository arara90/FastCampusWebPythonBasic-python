import requests
from bs4 import BeautifulSoup as bs
from konlpy.tag import Okt
from openpyxl import load_workbook as load

#크롬에서 태그 선택 오른쪽 > copy > copy selector
#https://www.fastcampus.co.kr/dev_online_introp/#text-block-17
URL = "https://music.naver.com/listen/top100.nhn?domain=TOTAL_V2"
save_dir = "C:\\Users\\ARA\\Documents\\FastCampusPythonBasic\\code11_Crawler\\rank.xlsx"

def get_tags():
    tags_list = []
    ul = []
    rq = requests.get(URL)
    print(rq.status_code)
    if rq.status_code==200:
        soup = bs(rq.content,'html.parser') 
        li_list = soup.find_all('td', class_='artist')

        for li in li_list:
            ul = li.select('a > span')
            for l in ul:
                tags_list.append(l.get_text())

        return tags_list


    else:
        raise Exception('200번(성공)이 아닙니다.')

def save_excel(rank):
    wb = load(save_dir)
    ws = wb.create_sheet('rank')
    idx=1
    try:
        for k,v in rank[:15]:
            ws['A' + str(idx)] = "{}({})".format(k,v)
            idx+=1
        wb.save(save_dir)

    except Exception as e:
        print(e)
    finally:
        wb.close()


def get_rank():
    ok=Okt()
    rank = {}
    tags_list = get_tags()
    #print(tags_list)
    for tag in tags_list:
        noun = ok.nouns(tag)
        for n in noun:
            if not (n in rank):
                rank[n] = 0
            rank[n] += 1

    rank = sorted(rank.items(),key = lambda x:x[1],reverse = True)
    save_excel(rank)
   

if __name__ == "__main__" :
    get_rank()

